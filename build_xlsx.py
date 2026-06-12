"""Generate the AI Agent Program plan as an Excel workbook.

Project tracker view: real calendar months and actual start/end dates.
Kickoff 17 June 2026, 5-month delivery window (finishes mid-November 2026).
"""
from calendar import monthrange
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- project calendar ----------
KICKOFF = date(2026, 6, 17)   # first working day of the programme


def add_months(d: date, n: int) -> date:
    m = d.month - 1 + n
    y = d.year + m // 12
    m = m % 12 + 1
    day = min(d.day, monthrange(y, m)[1])
    return date(y, m, day)


def phase_start(month_idx: int) -> date:
    """Actual start date for a 1-based project month index."""
    return add_months(KICKOFF, month_idx - 1)


def phase_end(month_idx: int) -> date:
    """Actual end date (inclusive) for a 1-based project month index."""
    return add_months(KICKOFF, month_idx) - timedelta(days=1)


def fmt(d: date) -> str:
    return d.strftime("%d %b %Y")


def fmt_short(d: date) -> str:
    return d.strftime("%d %b")



# ---------- styles ----------
NAVY = "232F3E"
ORANGE = "FF9900"
WHITE = "FFFFFF"
LIGHT = "F2F4F7"
GREY = "6B7480"
BAND = "EAF2F8"

PHASE_HEX = ["2E86C1", "28B463", "8E44AD", "E67E22", "16A085", "C0392B"]

thin = Side(style="thin", color="D5DBE0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(hexv):
    return PatternFill("solid", fgColor=hexv)


def style_cell(c, *, font_color="1B2430", bold=False, size=11, bg=None,
               align="left", wrap=False, valign="center", bordered=True):
    c.font = Font(color=font_color, bold=bold, size=size, name="Calibri")
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bordered:
        c.border = border


# (name, objective, start_month, dur_months, owner, aws, deliverables)
# Compressed 5-month plan - phases run in parallel to hit the deadline.
PHASES = [
    ("Phase 1 - Administrative AI Agent (MVP)",
     "Automate meeting admin: transcripts to minutes, actions, owners, due dates",
     1, 1, "AI/ML Lead",
     "Amazon Transcribe; Amazon Bedrock; Amazon S3; DynamoDB; AWS Lambda",
     "Meeting minutes; Action register; Summary report; Dashboard updates"),
    ("Phase 2 - Action Tracking Agent",
     "Monitor, remind, escalate and report on open / overdue actions",
     2, 1, "Automation Eng",
     "Amazon EventBridge; Amazon SES / SNS; DynamoDB; AWS Lambda",
     "Weekly action reports; Overdue alerts; Escalation notifications"),
    ("Phase 3 - Governance & Forum Agent",
     "Read SteerCo / CIO / Cyber / Risk / Audit outputs; detect themes & risks",
     2, 2, "Data Eng",
     "Amazon Bedrock; Bedrock Knowledge Bases; Amazon S3; AWS Lambda",
     "Governance dashboard; Executive reports; Risk insights"),
    ("Phase 4 - Reporting & Analytics Agent",
     "Consolidate reports, measure delivery, build CIO packs & trend analysis",
     3, 2, "BI Lead",
     "Amazon QuickSight; AWS Glue; Amazon Athena; Amazon S3",
     "QuickSight dashboards; Monthly reports; CIO packs; Trend analysis"),
    ("Phase 5 - Enterprise Knowledge Agent",
     "Searchable organisational memory; natural-language Q&A over all outputs",
     4, 1, "AI/ML Lead",
     "Bedrock Knowledge Bases; OpenSearch Serverless; Amazon Bedrock (RAG)",
     "Searchable knowledge base; Q&A assistant; Decision history"),
    ("Phase 6 - Central Control Tower (Orchestration)",
     'Connect all agents through a central platform - the "spine"',
     4, 2, "Solutions Architect",
     "Amazon Bedrock Agents; AWS Step Functions; API Gateway; Amazon Cognito",
     "Unified orchestration platform; Agent registry; Central console"),
]

PROJECT_START = phase_start(1)
PROJECT_END = phase_end(5)


def calendar_months(start: date, end: date):
    """Yield (year, month, first_day, last_day) for each calendar month in range."""
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        last = monthrange(y, m)[1]
        yield y, m, date(y, m, 1), date(y, m, last)
        m += 1
        if m > 12:
            m, y = 1, y + 1


CAL_MONTHS = list(calendar_months(PROJECT_START, PROJECT_END))



wb = Workbook()

# =====================================================================
# SHEET 1 - Roadmap (calendar Gantt)
# =====================================================================
ws = wb.active
ws.title = "Roadmap"
ws.sheet_view.showGridLines = False

n_cols = 3 + len(CAL_MONTHS)
last_col = get_column_letter(n_cols)

ws.merge_cells(f"A1:{last_col}1")
t = ws["A1"]
t.value = "Enterprise AI Agent Programme - Delivery Roadmap (AWS)"
style_cell(t, font_color=WHITE, bold=True, size=15, bg=NAVY, align="left", bordered=False)
ws.row_dimensions[1].height = 30

ws.merge_cells(f"A2:{last_col}2")
st = ws["A2"]
st.value = (f"Kickoff {fmt(PROJECT_START)}  -  Go-live {fmt(PROJECT_END)}"
            f"  (5-month delivery window)")
style_cell(st, font_color=ORANGE, bold=True, size=11, bg=NAVY, align="left", bordered=False)
ws.row_dimensions[2].height = 20

# Header row (row 4)
hdr_row = 4
headers = ["Phase", "Start", "End"] + [date(y, m, 1).strftime("%b %Y") for y, m, _, _ in CAL_MONTHS]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=hdr_row, column=j, value=h)
    align = "left" if j == 1 else "center"
    style_cell(c, font_color=WHITE, bold=True, size=10, bg=ORANGE, align=align)
ws.row_dimensions[hdr_row].height = 20

# Phase rows
for i, (name, obj, start, dur, owner, aws, deliv) in enumerate(PHASES):
    r = hdr_row + 1 + i
    s_date = phase_start(start)
    e_date = phase_end(start + dur - 1)
    style_cell(ws.cell(row=r, column=1, value=name), bold=True, size=10, align="left", wrap=True)
    # Real Excel date cells (not text) so they track like a project plan.
    cb = ws.cell(row=r, column=2, value=s_date)
    style_cell(cb, align="center", size=10)
    cb.number_format = "dd mmm yyyy"
    cc = ws.cell(row=r, column=3, value=e_date)
    style_cell(cc, align="center", size=10)
    cc.number_format = "dd mmm yyyy"
    first_active = True
    for k, (y, m, first, last) in enumerate(CAL_MONTHS):
        col = 4 + k
        cell = ws.cell(row=r, column=col)
        active = s_date <= last and e_date >= first
        style_cell(cell, bg=(PHASE_HEX[i] if active else LIGHT),
                   font_color=WHITE, bold=True, size=8, align="center", wrap=True)
        if active and first_active:
            cell.value = f"{fmt_short(s_date)} - {fmt_short(e_date)}"
            first_active = False
    ws.row_dimensions[r].height = 30



# Milestones row (actual target dates)
ms_row = hdr_row + 1 + len(PHASES) + 1
style_cell(ws.cell(row=ms_row, column=1, value="Key Milestones"),
           bold=True, size=10, bg=NAVY, font_color=WHITE)
style_cell(ws.cell(row=ms_row, column=2), bg=NAVY)
style_cell(ws.cell(row=ms_row, column=3), bg=NAVY)
MILESTONES = [
    (phase_end(1), "MVP live"),
    (phase_end(3), "Governance insights"),
    (phase_end(4), "Knowledge Q&A"),
    (phase_end(5), "Control Tower go-live"),
]
for k, (y, m, first, last) in enumerate(CAL_MONTHS):
    col = 4 + k
    cell = ws.cell(row=ms_row, column=col)
    label = next((f"{txt}\n({fmt_short(d)})" for d, txt in MILESTONES if first <= d <= last), None)
    if label:
        style_cell(cell, font_color=NAVY, bold=True, size=8, bg=ORANGE, align="center", wrap=True)
        cell.value = label
    else:
        style_cell(cell, bg=NAVY)
ws.row_dimensions[ms_row].height = 30

# Column widths
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 13
ws.column_dimensions["C"].width = 13
for k in range(len(CAL_MONTHS)):
    ws.column_dimensions[get_column_letter(4 + k)].width = 13

note_row = ms_row + 2
ws.cell(row=note_row, column=1,
        value=("Columns are calendar months. Shaded = phase active that month. "
               "Cross-cutting from kickoff: Security & IAM, cost governance, "
               "data privacy/PII, CI/CD, human-in-the-loop review."))
style_cell(ws.cell(row=note_row, column=1), font_color=GREY, size=9, bordered=False, align="left")
ws.freeze_panes = "D5"



# =====================================================================
# SHEET 2 - Phase Plan (detail with actual dates)
# =====================================================================
ws2 = wb.create_sheet("Phase Plan")
ws2.sheet_view.showGridLines = False
cols = ["Phase", "Start Date", "End Date", "Duration", "Owner",
        "Objective", "Core AWS Services", "Key Deliverables"]
widths = [30, 14, 14, 10, 18, 40, 38, 40]
for j, (h, w) in enumerate(zip(cols, widths), start=1):
    c = ws2.cell(row=1, column=j, value=h)
    style_cell(c, font_color=WHITE, bold=True, size=11, bg=NAVY, align="center", wrap=True)
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.row_dimensions[1].height = 26

for i, (name, obj, start, dur, owner, aws, deliv) in enumerate(PHASES):
    r = 2 + i
    band = BAND if i % 2 == 0 else WHITE
    s_date = phase_start(start)
    e_date = phase_end(start + dur - 1)
    vals = [name, s_date, e_date, f"{dur} mo", owner, obj,
            aws.replace("; ", "\n"), deliv.replace("; ", "\n")]
    for j, v in enumerate(vals, start=1):
        c = ws2.cell(row=r, column=j, value=v)
        style_cell(c, size=10, bg=band, wrap=True, valign="top", align="left")
    # Real date cells with a project-style date format.
    for jcol in (2, 3):
        ws2.cell(row=r, column=jcol).number_format = "dd mmm yyyy"
        ws2.cell(row=r, column=jcol).alignment = Alignment(horizontal="center", vertical="top")
    ws2.cell(row=r, column=1).font = Font(bold=True, size=10, color=PHASE_HEX[i], name="Calibri")
    ws2.row_dimensions[r].height = 72
ws2.freeze_panes = "A2"



# =====================================================================
# SHEET 3 - Detailed Tasks (with target date windows)
# =====================================================================
ws3 = wb.create_sheet("Detailed Tasks")
ws3.sheet_view.showGridLines = False
tcols = ["Phase", "Start", "End", "Workstream / Task", "AWS Service", "Output"]
twidths = [26, 13, 13, 44, 28, 36]
for j, (h, w) in enumerate(zip(tcols, twidths), start=1):
    c = ws3.cell(row=1, column=j, value=h)
    style_cell(c, font_color=WHITE, bold=True, size=11, bg=ORANGE, align="center")
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.row_dimensions[1].height = 22

# (phase_idx, start_month, end_month, task, aws_service, output)
TASKS = [
    (0, 1, 1, "Set up AWS landing zone, IAM roles, S3 buckets for raw inputs", "S3, IAM, Organizations", "Secure ingest foundation"),
    (0, 1, 1, "Transcribe recordings to text; ingest Teams transcripts & notes", "Amazon Transcribe, S3", "Normalised transcript store"),
    (0, 1, 1, "Extract actions/owners/due dates; minutes, summaries, register", "Amazon Bedrock, DynamoDB", "MVP live (minutes + register)"),
    (1, 2, 2, "Scheduled scan of open/overdue actions; reminder + escalation logic", "EventBridge, Lambda", "Automated reminders"),
    (1, 2, 2, "Notification channels and closure-evidence capture", "Amazon SES / SNS, S3", "Alerts & status reports"),
    (2, 2, 3, "Ingest SteerCo/CIO/Cyber/Risk/Audit outputs into knowledge base", "Bedrock Knowledge Bases, S3", "Governance corpus"),
    (2, 3, 3, "Theme, risk & dependency detection; executive summaries", "Amazon Bedrock, Lambda", "Risk insights & exec reports"),
    (3, 3, 3, "Build data lake / ETL for reporting; define KPIs & metrics", "AWS Glue, Athena, S3", "Curated reporting datasets"),
    (3, 3, 4, "Dashboards, monthly reports, CIO packs, trend analysis", "Amazon QuickSight", "Executive reporting packs"),
    (4, 4, 4, "Index all outputs; build vector store for semantic search", "OpenSearch Serverless, S3", "Searchable index"),
    (4, 4, 4, "RAG-based Q&A assistant over organisational memory", "Bedrock Knowledge Bases (RAG)", "Natural-language Q&A"),
    (5, 4, 4, "Design orchestration spine; agent registry & shared data contracts", "Step Functions, API Gateway", "Orchestration design"),
    (5, 5, 5, "Wire all agents together; central console & access control", "Bedrock Agents, Cognito", "Unified platform"),
    (5, 5, 5, "End-to-end testing, hardening, go-live & handover", "CloudWatch, Step Functions", "Production Control Tower"),
]
for i, (pidx, sm, em, task, svc, out) in enumerate(TASKS):
    r = 2 + i
    band = BAND if pidx % 2 == 0 else WHITE
    row_vals = [PHASES[pidx][0].split(" - ")[0], phase_start(sm),
                phase_end(em), task, svc, out]
    for j, v in enumerate(row_vals, start=1):
        c = ws3.cell(row=r, column=j, value=v)
        style_cell(c, size=10, bg=band, wrap=True, valign="center", align="left")
    for jcol in (2, 3):
        ws3.cell(row=r, column=jcol).number_format = "dd mmm yyyy"
        ws3.cell(row=r, column=jcol).alignment = Alignment(horizontal="center", vertical="center")
    ws3.cell(row=r, column=1).font = Font(bold=True, size=9, color=PHASE_HEX[pidx], name="Calibri")
    ws3.row_dimensions[r].height = 30
ws3.freeze_panes = "A2"

import os
_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AI_Agent_Programme_Plan.xlsx")
wb.save(_OUT)
print("XLSX saved:", _OUT)
print("Sheets:", wb.sheetnames)
print("Calendar months:", [date(y, m, 1).strftime("%b %Y") for y, m, _, _ in CAL_MONTHS])
